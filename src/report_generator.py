# src/report_generator.py
"""
Report Generator - Creates Excel reports matching client template exactly
"""

import os
from datetime import datetime
from typing import List, Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from config_manager import ConfigManager


class ReportGenerator:
    """Generates Excel reports from analysis results"""
    
    def __init__(self, config_manager: ConfigManager):
        self.config = config_manager
        self.campaign_name = config_manager.get_campaign_name()
        self.target_brand = config_manager.get_target_brand()
    
    def create_report(self, platform_results: Dict[str, List[Dict[str, Any]]],
                     output_dir: str = "output") -> str:
        """Create Excel report from results - one sheet per platform"""

        # Create output directory
        os.makedirs(output_dir, exist_ok=True)

        # Create workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Create sheet for each platform
        for platform_name, results in platform_results.items():
            self._create_platform_sheet(wb, platform_name, results)

        # Add campaign info sheet
        self._create_campaign_info_sheet(wb)

        # Generate filename with DEBUG logging
        timestamp = datetime.now().strftime('%Y%m%d_%H%M')

        # DEBUG: Log each step of filename generation
        print(f"DEBUG FILENAME: self.campaign_name = '{self.campaign_name}'")
        campaign_name_cleaned = self.campaign_name.replace(' ', '_')
        print(f"DEBUG FILENAME: campaign_name_cleaned = '{campaign_name_cleaned}'")
        print(f"DEBUG FILENAME: timestamp = '{timestamp}'")

        filename = f"{output_dir}/{campaign_name_cleaned}_{timestamp}.xlsx"
        print(f"DEBUG FILENAME: Final filename = '{filename}'")

        # Save
        wb.save(filename)
        print(f"DEBUG FILENAME: File saved to '{filename}'")

        # Verify the file exists
        if os.path.exists(filename):
            print(f"DEBUG FILENAME: Verified file exists at '{filename}'")
        else:
            print(f"DEBUG FILENAME: WARNING - File NOT found at '{filename}'")

        return filename
    
    def _create_platform_sheet(self, wb: Workbook, platform_name: str, 
                               results: List[Dict[str, Any]]):
        """Create sheet for a specific platform matching client template"""
        ws = wb.create_sheet(platform_name)
        
        # Get custom keyword category names from config
        keyword_categories = list(self.config.get_keywords().keys())
        
        # Build headers dynamically with custom category names
        headers = [
            "Query",
            "Query brand (they want to audit)",
            "Organic competitor 1",
            "Likelihood 1",
            "Organic competitor 2",
            "Likelihood 2",
            "Organic competitor 3",
            "Likelihood 3",
            "Position",
            "Competitors Mentioned",
            "Source(s) Cited",
            "Recency of primary source",
        ]
        
        # Add custom keyword category headers
        for category_name in keyword_categories:
            headers.append(f"Inclusion of key messages - {category_name}")
        
        # Add remaining headers
        headers.extend([
            "Tone (POS/NEG/NEU)",
            "Style mentioned",
            "Time and date of query run"
        ])
        
        ws.append(headers)
        
        # Style headers
        self._style_header_row(ws, 1)
        
        # Add data rows
        for result in results:
            row_data = self._format_platform_row(result, keyword_categories)
            ws.append(row_data)
        
        # Format columns
        self._format_columns(ws, len(keyword_categories))
    
    def _format_platform_row(self, result: Dict[str, Any], keyword_categories: List[str]) -> List[Any]:
        """Format result into row matching client template"""
        # 1. Query
        query = result['query']
        
        # 2. Query brand - show TARGET BRAND from EITHER query_brands OR organic_competitors
        query_brands = result['query_brands']
        organic = result['organic_competitors']
        total_runs = result['total_runs']
        target_brand_lower = self.target_brand.lower()
        
        # First check query_brands (brands mentioned in the query)
        target_brand_stats = None
        for brand, likelihood in query_brands:
            if brand.lower() == target_brand_lower:
                target_brand_stats = (brand, likelihood)
                break
        
        # If not in query_brands, check organic_competitors (brands found in responses)
        if not target_brand_stats:
            for brand, count in organic:
                if brand.lower() == target_brand_lower:
                    likelihood = round((count / total_runs) * 100)
                    target_brand_stats = (brand, likelihood)
                    break
        
        # Format the target brand string
        if target_brand_stats:
            query_brand_str = f"{target_brand_stats[0]} ({target_brand_stats[1]}%)"
        else:
            query_brand_str = f"{self.target_brand} (0%)"
        
        # 3-8. Organic competitors (excluding target brand)
        organic_filtered = [(brand, count) for brand, count in organic 
                           if brand.lower() != target_brand_lower]
        
        organic_1 = organic_filtered[0][0] if len(organic_filtered) > 0 else ""
        organic_1_pct = f"{round((organic_filtered[0][1]/total_runs)*100)}%" if len(organic_filtered) > 0 else ""
        
        organic_2 = organic_filtered[1][0] if len(organic_filtered) > 1 else ""
        organic_2_pct = f"{round((organic_filtered[1][1]/total_runs)*100)}%" if len(organic_filtered) > 1 else ""
        
        organic_3 = organic_filtered[2][0] if len(organic_filtered) > 2 else ""
        organic_3_pct = f"{round((organic_filtered[2][1]/total_runs)*100)}%" if len(organic_filtered) > 2 else ""
        
        # 9. Position
        position = result.get('position', "")
        
        # 10. Competitors Mentioned
        all_competitors = []
        for brand, _ in query_brands:
            if brand.lower() != target_brand_lower:
                all_competitors.append(brand)
        for brand, _ in organic:
            if brand.lower() != target_brand_lower and brand not in all_competitors:
                all_competitors.append(brand)
        competitors_mentioned = ", ".join(all_competitors) if all_competitors else ""
        
        # 11. Sources
        sources = result['sources']
        sources_str = "; ".join([f"{domain} ({count}x)" for domain, count in sources[:5]]) \
                     if sources else ""
        
        # 12. Recency
        recency = result.get('source_recency', "")
        
        # 13-N. Key messages - USE ACTUAL CATEGORY NAMES (case-sensitive)
        key_messages = result['key_messages']
        keyword_percentages = []
        for category_name in keyword_categories:
            pct = key_messages.get(category_name, 0)
            keyword_percentages.append(f"{pct}%")
        
        # N+1. Tone
        tone = result.get('tone', "N")
        
        # N+2. Models
        models = result['models']
        models_str = ", ".join([model for model, count in models[:10]]) if models else ""
        
        # N+3. Timestamp
        timestamp = result.get('timestamp', datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        
        # Build row
        row = [
            query,
            query_brand_str,
            organic_1, organic_1_pct,
            organic_2, organic_2_pct,
            organic_3, organic_3_pct,
            position,
            competitors_mentioned,
            sources_str,
            recency,
        ]
        
        # Add keyword percentages dynamically
        row.extend(keyword_percentages)
        
        # Add remaining fields
        row.extend([tone, models_str, timestamp])
        
        return row
    
    def _create_campaign_info_sheet(self, wb: Workbook):
        """Create campaign info sheet"""
        ws = wb.create_sheet("Campaign Info")
        
        info_data = [
            ["Campaign Information", ""],
            ["", ""],
            ["Campaign Name", self.config.get_campaign_name()],
            ["Target Brand", self.config.get_target_brand()],
            ["Runs per Query (per platform)", self.config.get_runs_per_query()],
            ["Date Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            ["", ""],
            ["Platforms", ""],
        ]
        
        for platform in self.config.get_enabled_platforms():
            info_data.append(["", f"✓ {platform.title()}"])
        
        info_data.append(["", ""])
        info_data.append(["Competitors", ""])
        
        for competitor in self.config.get_competitors():
            info_data.append(["", competitor])
        
        for row in info_data:
            ws.append(row)
        
        ws['A1'].font = Font(bold=True, size=14)
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 40
    
    def _style_header_row(self, ws, row_num: int):
        """Apply styling to header row"""
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[row_num]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    def _format_columns(self, ws, num_keyword_categories: int):
        """Format column widths"""
        base_columns = {
            'A': 50,  # Query
            'B': 25,  # Query brand
            'C': 20,  # Organic 1
            'D': 12,  # Likelihood 1
            'E': 20,  # Organic 2
            'F': 12,  # Likelihood 2
            'G': 20,  # Organic 3
            'H': 12,  # Likelihood 3
            'I': 12,  # Position
            'J': 30,  # Competitors Mentioned
            'K': 45,  # Sources
            'L': 20,  # Recency
        }
        
        for col, width in base_columns.items():
            ws.column_dimensions[col].width = width
        
        # Keyword columns (M, N, O, P, etc.) - dynamic based on number of categories
        keyword_start_col = 13  # Column M
        for i in range(num_keyword_categories):
            col_letter = chr(ord('M') + i)
            ws.column_dimensions[col_letter].width = 15
        
        # Remaining columns after keywords
        tone_col = chr(ord('M') + num_keyword_categories)
        models_col = chr(ord('M') + num_keyword_categories + 1)
        timestamp_col = chr(ord('M') + num_keyword_categories + 2)
        
        ws.column_dimensions[tone_col].width = 8
        ws.column_dimensions[models_col].width = 40
        ws.column_dimensions[timestamp_col].width = 20
        
        # Cell alignment
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical='top', wrap_text=True)